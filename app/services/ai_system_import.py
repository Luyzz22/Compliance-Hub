from __future__ import annotations

import csv
import io
import json
import re
import zipfile
from pathlib import PurePosixPath
from uuid import uuid4

from openpyxl import load_workbook
from pydantic import BaseModel, ValidationError

from app.ai_system_models import (
    AIActCategory,
    AIImportResult,
    AIImportRowError,
    AISystem,
    AISystemCreate,
    AISystemCriticality,
    AISystemRiskLevel,
    AISystemUpdate,
    DataSensitivity,
)
from app.policy_service import evaluate_policies_for_ai_system
from app.repositories.ai_systems import AISystemRepository
from app.repositories.audit import AuditRepository
from app.repositories.audit_logs import AuditLogRepository
from app.repositories.policies import PolicyRepository
from app.repositories.violations import ViolationRepository
from app.security import AuthContext

ALLOWED_FIELDS = frozenset(
    {
        "id",
        "name",
        "description",
        "business_unit",
        "risk_level",
        "ai_act_category",
        "owner_email",
        "criticality",
        "data_sensitivity",
        "has_incident_runbook",
        "has_supplier_risk_register",
        "has_backup_runbook",
        "gdpr_dpia_required",
    }
)

HEADER_SYNONYMS: dict[str, str] = {
    "system_id": "id",
    "uuid": "id",
    "risiko": "risk_level",
    "risikostufe": "risk_level",
    "risk": "risk_level",
    "ai_category": "ai_act_category",
    "eu_ai_act_category": "ai_act_category",
    "kategorie": "ai_act_category",
    "bu": "business_unit",
    "geschaeftseinheit": "business_unit",
    "abteilung": "business_unit",
    "businessunit": "business_unit",
    "owner": "owner_email",
    "e_mail": "owner_email",
    "dpia": "gdpr_dpia_required",
    "gdpr_dpia": "gdpr_dpia_required",
    "incident_runbook": "has_incident_runbook",
    "supplier_risk": "has_supplier_risk_register",
    "supplier_register": "has_supplier_risk_register",
    "backup_runbook": "has_backup_runbook",
}

MAX_IMPORT_UPLOAD_BYTES = 5 * 1024 * 1024
MAX_IMPORT_ROWS = 10_000
MAX_IMPORT_COLUMNS = 64
MAX_IMPORT_CELL_CHARS = 10_000
MAX_XLSX_ARCHIVE_ENTRIES = 256
MAX_XLSX_EXPANDED_BYTES = 40 * 1024 * 1024
MAX_XLSX_ENTRY_BYTES = 20 * 1024 * 1024
MAX_XLSX_COMPRESSION_RATIO = 150


def _normalize_header_key(cell: str) -> str:
    s = cell.strip().lower().replace("\ufeff", "")
    s = re.sub(r"[\s\-]+", "_", s)
    return s


def _resolve_field(normalized_header: str) -> str | None:
    key = HEADER_SYNONYMS.get(normalized_header, normalized_header)
    return key if key in ALLOWED_FIELDS else None


def _parse_bool(raw: str | None, default: bool = False) -> bool:
    if raw is None or not str(raw).strip():
        return default
    s = str(raw).strip().lower()
    if s in ("1", "true", "t", "yes", "y", "ja", "j", "wahr", "x", "✓", "ok"):
        return True
    if s in ("0", "false", "f", "no", "n", "nein", "-", ""):
        return False
    raise ValueError(f"Ungültiger Boolescher Wert: {raw!r}")


def _parse_risk_level(raw: str) -> AISystemRiskLevel:
    s = raw.strip().lower().replace(" ", "_").replace("-", "_")
    mapping: dict[str, AISystemRiskLevel] = {
        "low": AISystemRiskLevel.low,
        "niedrig": AISystemRiskLevel.low,
        "gering": AISystemRiskLevel.low,
        "limited": AISystemRiskLevel.limited,
        "begrenzt": AISystemRiskLevel.limited,
        "high": AISystemRiskLevel.high,
        "hoch": AISystemRiskLevel.high,
        "elevated": AISystemRiskLevel.high,
        "unacceptable": AISystemRiskLevel.unacceptable,
        "unakzeptabel": AISystemRiskLevel.unacceptable,
        "inakzeptabel": AISystemRiskLevel.unacceptable,
    }
    if s in mapping:
        return mapping[s]
    try:
        return AISystemRiskLevel(s)
    except ValueError as exc:
        raise ValueError(f"Ungültiges risk_level: {raw!r}") from exc


def _parse_ai_act_category(raw: str) -> AIActCategory:
    s = raw.strip().lower().replace(" ", "_").replace("-", "_")
    aliases: dict[str, AIActCategory] = {
        "highrisk": AIActCategory.high_risk,
        "high_risk": AIActCategory.high_risk,
        "hohes_risiko": AIActCategory.high_risk,
        "limitedrisk": AIActCategory.limited_risk,
        "limited_risk": AIActCategory.limited_risk,
        "begrenztes_risiko": AIActCategory.limited_risk,
        "minimalrisk": AIActCategory.minimal_risk,
        "minimal_risk": AIActCategory.minimal_risk,
        "minimales_risiko": AIActCategory.minimal_risk,
        "prohibited": AIActCategory.prohibited,
        "verboten": AIActCategory.prohibited,
    }
    if s in aliases:
        return aliases[s]
    try:
        return AIActCategory(s)
    except ValueError as exc:
        raise ValueError(f"Ungültige ai_act_category: {raw!r}") from exc


def _parse_criticality(raw: str | None) -> AISystemCriticality:
    if raw is None or not str(raw).strip():
        return AISystemCriticality.medium
    s = str(raw).strip().lower().replace(" ", "_").replace("-", "_")
    aliases: dict[str, AISystemCriticality] = {
        "low": AISystemCriticality.low,
        "niedrig": AISystemCriticality.low,
        "medium": AISystemCriticality.medium,
        "mittel": AISystemCriticality.medium,
        "high": AISystemCriticality.high,
        "hoch": AISystemCriticality.high,
        "very_high": AISystemCriticality.very_high,
        "veryhigh": AISystemCriticality.very_high,
        "sehr_hoch": AISystemCriticality.very_high,
        "sehrhoch": AISystemCriticality.very_high,
    }
    if s in aliases:
        return aliases[s]
    try:
        return AISystemCriticality(s)
    except ValueError as exc:
        raise ValueError(f"Ungültige criticality: {raw!r}") from exc


def _parse_data_sensitivity(raw: str | None) -> DataSensitivity:
    if raw is None or not str(raw).strip():
        return DataSensitivity.internal
    s = str(raw).strip().lower().replace(" ", "_").replace("-", "_")
    aliases: dict[str, DataSensitivity] = {
        "public": DataSensitivity.public,
        "oeffentlich": DataSensitivity.public,
        "internal": DataSensitivity.internal,
        "intern": DataSensitivity.internal,
        "confidential": DataSensitivity.confidential,
        "vertraulich": DataSensitivity.confidential,
        "restricted": DataSensitivity.restricted,
        "streng_vertraulich": DataSensitivity.restricted,
    }
    if s in aliases:
        return aliases[s]
    try:
        return DataSensitivity(s)
    except ValueError as exc:
        raise ValueError(f"Ungültige data_sensitivity: {raw!r}") from exc


def _row_to_create_payload(row: dict[str, str], resolved_id: str) -> AISystemCreate:
    name = (row.get("name") or "").strip()
    description = (row.get("description") or "").strip()
    business_unit = (row.get("business_unit") or "").strip()
    if not name:
        raise ValueError("Pflichtfeld name ist leer")
    if not description:
        raise ValueError("Pflichtfeld description ist leer")
    if not business_unit:
        raise ValueError("Pflichtfeld business_unit ist leer")
    rl_raw = (row.get("risk_level") or "").strip()
    ac_raw = (row.get("ai_act_category") or "").strip()
    if not rl_raw:
        raise ValueError("Pflichtfeld risk_level ist leer")
    if not ac_raw:
        raise ValueError("Pflichtfeld ai_act_category ist leer")
    owner = (row.get("owner_email") or "").strip() or None
    return AISystemCreate(
        id=resolved_id,
        name=name,
        description=description,
        business_unit=business_unit,
        risk_level=_parse_risk_level(rl_raw),
        ai_act_category=_parse_ai_act_category(ac_raw),
        gdpr_dpia_required=_parse_bool(row.get("gdpr_dpia_required"), False),
        owner_email=owner,
        criticality=_parse_criticality(row.get("criticality")),
        data_sensitivity=_parse_data_sensitivity(row.get("data_sensitivity")),
        has_incident_runbook=_parse_bool(row.get("has_incident_runbook"), False),
        has_supplier_risk_register=_parse_bool(row.get("has_supplier_risk_register"), False),
        has_backup_runbook=_parse_bool(row.get("has_backup_runbook"), False),
    )


def _create_to_update(create: AISystemCreate) -> AISystemUpdate:
    return AISystemUpdate(
        name=create.name,
        description=create.description,
        business_unit=create.business_unit,
        risk_level=create.risk_level,
        ai_act_category=create.ai_act_category,
        gdpr_dpia_required=create.gdpr_dpia_required,
        owner_email=create.owner_email,
        criticality=create.criticality,
        data_sensitivity=create.data_sensitivity,
        has_incident_runbook=create.has_incident_runbook,
        has_supplier_risk_register=create.has_supplier_risk_register,
        has_backup_runbook=create.has_backup_runbook,
    )


def _model_to_json(model: BaseModel) -> str:
    return json.dumps(model.model_dump(mode="json"), default=str)


def _bounded_cell(value: object) -> str:
    rendered = "" if value is None else str(value).strip()
    if len(rendered) > MAX_IMPORT_CELL_CHARS:
        raise ValueError(
            f"Zellen dürfen höchstens {MAX_IMPORT_CELL_CHARS} Zeichen enthalten",
        )
    return rendered


def _iter_csv_rows(data: bytes) -> list[tuple[int, dict[str, str]]]:
    text = data.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text), strict=True)
    header_cells = next(reader, None)
    if header_cells is None:
        return []
    if len(header_cells) > MAX_IMPORT_COLUMNS:
        raise ValueError(f"CSV darf höchstens {MAX_IMPORT_COLUMNS} Spalten enthalten")
    canonical = [_resolve_field(_normalize_header_key(h or "")) for h in header_cells]
    out: list[tuple[int, dict[str, str]]] = []
    for row_idx, row in enumerate(reader, start=2):
        if row_idx > MAX_IMPORT_ROWS + 1:
            raise ValueError(f"Import darf höchstens {MAX_IMPORT_ROWS} Datenzeilen enthalten")
        if len(row) > MAX_IMPORT_COLUMNS:
            raise ValueError(f"CSV darf höchstens {MAX_IMPORT_COLUMNS} Spalten enthalten")
        d: dict[str, str] = {}
        for j, key in enumerate(canonical):
            if key is None:
                continue
            val = (row[j] if j < len(row) else "") or ""
            d[key] = _bounded_cell(val)
        if not any(v for v in d.values()):
            continue
        out.append((row_idx, d))
    return out


def _inspect_xlsx_archive(data: bytes) -> None:
    try:
        archive = zipfile.ZipFile(io.BytesIO(data))
    except zipfile.BadZipFile as exc:
        raise ValueError("Ungültiges XLSX-Archiv") from exc

    with archive:
        entries = archive.infolist()
        if len(entries) > MAX_XLSX_ARCHIVE_ENTRIES:
            raise ValueError("XLSX enthält zu viele Archiveinträge")
        expanded_total = 0
        for entry in entries:
            path = PurePosixPath(entry.filename)
            if path.is_absolute() or ".." in path.parts:
                raise ValueError("XLSX enthält einen unsicheren Archivpfad")
            if entry.flag_bits & 0x1:
                raise ValueError("Verschlüsselte XLSX-Archive werden nicht unterstützt")
            if entry.file_size > MAX_XLSX_ENTRY_BYTES:
                raise ValueError("XLSX enthält einen zu großen Archiveintrag")
            expanded_total += entry.file_size
            if expanded_total > MAX_XLSX_EXPANDED_BYTES:
                raise ValueError("XLSX überschreitet das erlaubte entpackte Datenvolumen")
            if entry.file_size and entry.compress_size == 0:
                raise ValueError("XLSX enthält einen ungültigen Kompressionswert")
            compression_ratio = entry.file_size / entry.compress_size if entry.compress_size else 0
            if compression_ratio > MAX_XLSX_COMPRESSION_RATIO:
                raise ValueError("XLSX weist eine unzulässige Kompressionsrate auf")


def _iter_xlsx_rows(data: bytes) -> list[tuple[int, dict[str, str]]]:
    _inspect_xlsx_archive(data)
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        if len(wb.sheetnames) > 10:
            raise ValueError("XLSX darf höchstens 10 Tabellenblätter enthalten")
        ws = wb.active
        if ws.max_column > MAX_IMPORT_COLUMNS:
            raise ValueError(f"XLSX darf höchstens {MAX_IMPORT_COLUMNS} Spalten enthalten")
        if ws.max_row > MAX_IMPORT_ROWS + 1:
            raise ValueError(f"Import darf höchstens {MAX_IMPORT_ROWS} Datenzeilen enthalten")
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return []
        header_cells = [_bounded_cell(c) for c in header_row]
        canonical = [_resolve_field(_normalize_header_key(h)) for h in header_cells]
        out: list[tuple[int, dict[str, str]]] = []
        for row_idx, row in enumerate(rows_iter, start=2):
            if row_idx > MAX_IMPORT_ROWS + 1:
                raise ValueError(f"Import darf höchstens {MAX_IMPORT_ROWS} Datenzeilen enthalten")
            if row is None:
                continue
            d: dict[str, str] = {}
            for j, key in enumerate(canonical):
                if key is None:
                    continue
                cell = row[j] if j < len(row) else None
                d[key] = _bounded_cell(cell)
            if not any(v for v in d.values()):
                continue
            out.append((row_idx, d))
        return out
    finally:
        wb.close()


def _parse_table_rows(filename: str, data: bytes) -> list[tuple[int, dict[str, str]]]:
    name = filename.lower()
    if name.endswith(".xlsx"):
        return _iter_xlsx_rows(data)
    if name.endswith(".csv"):
        return _iter_csv_rows(data)
    raise ValueError("Nur CSV- und XLSX-Dateien werden unterstützt")


def _after_create(
    tenant_id: str,
    auth: AuthContext,
    created: AISystem,
    audit_log_repo: AuditLogRepository,
    audit_event_repo: AuditRepository,
    policy_repo: PolicyRepository,
    violation_repo: ViolationRepository,
) -> None:
    policy_repo.ensure_default_policy_rules(tenant_id)
    audit_log_repo.record_event(
        tenant_id=tenant_id,
        actor="system",
        action="create_ai_system",
        entity_type="AISystem",
        entity_id=created.id,
        before=None,
        after=_model_to_json(created),
    )
    audit_event_repo.log_event(
        tenant_id=tenant_id,
        actor_type="api_key",
        actor_id=auth.actor_id,
        entity_type="ai_system",
        entity_id=created.id,
        action="created",
        metadata={"status": created.status.value},
    )
    evaluate_policies_for_ai_system(
        tenant_id=tenant_id,
        ai_system=created,
        policy_repository=policy_repo,
        violation_repository=violation_repo,
        audit_repository=audit_event_repo,
        actor_type="api_key",
        actor_id=auth.actor_id,
    )


def _after_update(
    tenant_id: str,
    auth: AuthContext,
    updated: AISystem,
    audit_event_repo: AuditRepository,
    policy_repo: PolicyRepository,
    violation_repo: ViolationRepository,
) -> None:
    audit_event_repo.log_event(
        tenant_id=tenant_id,
        actor_type="api_key",
        actor_id=auth.actor_id,
        entity_type="ai_system",
        entity_id=updated.id,
        action="updated",
        metadata={"status": updated.status.value},
    )
    evaluate_policies_for_ai_system(
        tenant_id=tenant_id,
        ai_system=updated,
        policy_repository=policy_repo,
        violation_repository=violation_repo,
        audit_repository=audit_event_repo,
        actor_type="api_key",
        actor_id=auth.actor_id,
    )


def import_ai_systems_from_file(
    *,
    tenant_id: str,
    auth: AuthContext,
    filename: str,
    data: bytes,
    repository: AISystemRepository,
    audit_log_repo: AuditLogRepository,
    audit_event_repo: AuditRepository,
    policy_repo: PolicyRepository,
    violation_repo: ViolationRepository,
) -> AIImportResult:
    errors: list[AIImportRowError] = []
    imported = 0

    try:
        parsed = _parse_table_rows(filename, data)
    except Exception as exc:  # noqa: BLE001 — tolerant catch for malformed XLSX/CSV
        return AIImportResult(
            total_rows=0,
            imported_count=0,
            failed_count=1,
            errors=[
                AIImportRowError(row_number=0, message=f"Datei konnte nicht gelesen werden: {exc}")
            ],
        )

    total = len(parsed)

    for row_number, row in parsed:
        try:
            explicit_id = (row.get("id") or "").strip()
            name = (row.get("name") or "").strip()

            if explicit_id:
                target_id = explicit_id
                existing = repository.get_by_id(tenant_id, target_id)
                payload = _row_to_create_payload(row, target_id)
                if existing is None:
                    created = repository.create(tenant_id, payload)
                    _after_create(
                        tenant_id,
                        auth,
                        created,
                        audit_log_repo,
                        audit_event_repo,
                        policy_repo,
                        violation_repo,
                    )
                else:
                    updated = repository.update(
                        tenant_id,
                        target_id,
                        _create_to_update(payload),
                    )
                    _after_update(
                        tenant_id,
                        auth,
                        updated,
                        audit_event_repo,
                        policy_repo,
                        violation_repo,
                    )
            else:
                if not name:
                    raise ValueError("Ohne id ist name Pflicht")
                existing = repository.get_by_name(tenant_id, name)
                new_id = str(uuid4()) if existing is None else existing.id
                payload = _row_to_create_payload(row, new_id)
                if existing is None:
                    created = repository.create(tenant_id, payload)
                    _after_create(
                        tenant_id,
                        auth,
                        created,
                        audit_log_repo,
                        audit_event_repo,
                        policy_repo,
                        violation_repo,
                    )
                else:
                    updated = repository.update(
                        tenant_id,
                        existing.id,
                        _create_to_update(payload),
                    )
                    _after_update(
                        tenant_id,
                        auth,
                        updated,
                        audit_event_repo,
                        policy_repo,
                        violation_repo,
                    )
            imported += 1
        except (ValueError, ValidationError) as exc:
            msg = str(exc)
            if isinstance(exc, ValidationError):
                msg = "; ".join(f"{e['loc']}: {e['msg']}" for e in exc.errors())
            errors.append(AIImportRowError(row_number=row_number, message=msg))

    return AIImportResult(
        total_rows=total,
        imported_count=imported,
        failed_count=len(errors),
        errors=errors,
    )
